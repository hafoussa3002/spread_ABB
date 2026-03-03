from __future__ import annotations

import logging
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook


LOGGER = logging.getLogger(__name__)


@dataclass
class FilterConfig:
    issue_start: date = date(2023, 1, 3)
    issue_end: date = date(2025, 12, 31)
    maturity_start: date = date(2023, 1, 1)
    maturity_end: date = date(2035, 12, 31)
    residual_min_days: int = 1
    residual_max_days: int = 1830


def _ensure_required_columns(df: pd.DataFrame) -> None:
    required = [
        "ISSUEDT",
        "MATURITYDT_L",
        "INSTRCTGRY",
        "ENGLONGNAME",
        "ENGPREFERREDNAME",
    ]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Colonnes manquantes: {missing}")


def load_sheet(path: str | Path, sheet_name: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
    _ensure_required_columns(df)
    return df


def filter_cd_rows(df: pd.DataFrame, cfg: Optional[FilterConfig] = None) -> pd.Series:
    cfg = cfg or FilterConfig()
    dff = df.copy()
    dff["ISSUEDT"] = pd.to_datetime(dff["ISSUEDT"], errors="coerce").dt.date
    dff["MATURITYDT_L"] = pd.to_datetime(dff["MATURITYDT_L"], errors="coerce").dt.date

    name_mix = (
        dff["ENGLONGNAME"].fillna("").astype(str)
        + " "
        + dff["ENGPREFERREDNAME"].fillna("").astype(str)
    )
    mask_cd = (
        dff["INSTRCTGRY"].fillna("").astype(str).str.upper().eq("TCN")
        & name_mix.str.contains("CD", case=False, regex=False)
    )

    mask_dates = (
        dff["ISSUEDT"].between(cfg.issue_start, cfg.issue_end)
        & dff["MATURITYDT_L"].between(cfg.maturity_start, cfg.maturity_end)
    )

    residual_days = (pd.to_datetime(dff["MATURITYDT_L"]) - pd.to_datetime(dff["ISSUEDT"])).dt.days
    mask_residual = residual_days.between(cfg.residual_min_days, cfg.residual_max_days)

    mask = mask_cd & mask_dates & mask_residual
    LOGGER.info("Lignes totales: %d | lignes CD retenues: %d", len(df), int(mask.sum()))
    return mask


def write_rates_to_excel(
    input_path: str | Path,
    output_path: str | Path,
    sheet_name: str,
    rates_by_df_index: dict[int, float | None],
    taux_col_name: str = "Taux BDT",
    taux_number_format: str = "0.000%",
) -> None:
    wb = load_workbook(input_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Feuille introuvable: {sheet_name}")
    ws = wb[sheet_name]

    headers = {str(ws.cell(row=1, column=c).value).strip(): c for c in range(1, ws.max_column + 1)}
    taux_col = headers.get(taux_col_name)
    if taux_col is None:
        taux_col = ws.max_column + 1
        ws.cell(row=1, column=taux_col).value = taux_col_name
        LOGGER.info("Colonne '%s' créée en position %d", taux_col_name, taux_col)

    for idx, val in rates_by_df_index.items():
        excel_row = int(idx) + 2
        cell = ws.cell(row=excel_row, column=taux_col)
        cell.value = None if val is None else float(val)
        if val is not None:
            cell.number_format = taux_number_format

    wb.save(output_path)
    LOGGER.info("Fichier sauvegardé: %s", Path(output_path).resolve())
