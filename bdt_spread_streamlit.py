from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from pathlib import Path
import re
from concurrent.futures import ThreadPoolExecutor
from typing import Optional

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from bam_curve_fetcher import BamCurveFetcher
from vba_equivalent_rates import calcul_taux


REQUIRED_COLUMNS = {
    "ENGLONGNAME",
    "ISSUEDT",
    "MATURITYDT_L",
}

INSTRUMENTS = ["CD", "BSF"]
APP_USER = "spreadABB"
APP_CODE = "albarid2026"


def _resolve_logo_path() -> Optional[Path]:
    candidates = [
        Path(r"C:\Users\tahaf\anaconda_projects\assets\ALBARID.png"),
        Path("assets/ALBARID.png"),
    ]
    for p in candidates:
        if p.exists():
            return p
    return None


def _inject_theme_css() -> None:
    st.markdown(
        """
        <style>
        .stApp {
            background: radial-gradient(circle at 25% 20%, #7c5f51 0%, #5a443b 38%, #3b2d2a 100%);
            color: #f6f1e9;
        }
        [data-testid="stHeader"] {
            background: transparent !important;
            height: 0 !important;
        }
        [data-testid="stToolbar"] {
            top: 0.25rem !important;
        }
        .stSidebar {
            background: linear-gradient(180deg, #6b544a 0%, #4d3a33 55%, #372b27 100%);
        }
        .stSidebar h1, .stSidebar h2, .stSidebar h3, .stSidebar p, .stSidebar label {
            color: #e9d7b7 !important;
        }
        .block-container {
            padding-top: 1.3rem;
        }
        .abb-card {
            border: 1px solid rgba(255,255,255,0.22);
            background: rgba(255,255,255,0.10);
            border-radius: 16px;
            padding: 18px;
            margin-bottom: 14px;
            backdrop-filter: blur(3px);
        }
        .abb-title {
            font-size: 44px;
            font-weight: 700;
            margin-bottom: 6px;
        }
        .abb-subtitle {
            font-size: 20px;
            opacity: 0.95;
            margin-bottom: 18px;
        }
        .app-card {
            background: #f4f0e7;
            border-left: 6px solid #e3c100;
            border-radius: 14px;
            padding: 14px 16px;
            color: #16253d;
            margin-bottom: 14px;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def _render_login(logo_path: Optional[Path]) -> None:
    st.markdown(
        "<div class='abb-card'><div class='abb-title'>Plateforme de Spread</div>"
        "<div class='abb-subtitle'><b>Al Barid Bank</b><br/>Connexion securisee a la plateforme interne</div></div>",
        unsafe_allow_html=True,
    )
    c1, c2, c3 = st.columns([1.2, 1, 1.2])
    with c2:
        if logo_path is not None:
            st.image(str(logo_path), width=90)
        with st.form("login_form", clear_on_submit=False):
            username = st.text_input("Username")
            code = st.text_input("Code", type="password")
            submit = st.form_submit_button("Se connecter", use_container_width=True)
        if submit:
            if username == APP_USER and code == APP_CODE:
                st.session_state.auth_ok = True
                st.rerun()
            else:
                st.error("Identifiants invalides.")


def _render_top_banner(logo_path: Optional[Path]) -> None:
    c1, c2 = st.columns([0.12, 0.88])
    with c1:
        if logo_path is not None:
            st.image(str(logo_path), width=86)
    with c2:
        st.markdown(
            "<div class='app-card'><h3 style='margin:0'>Plateforme de Spread</h3>"
            "<div>Plateforme interne de pilotage</div></div>",
            unsafe_allow_html=True,
        )


def _norm_col(name: object) -> str:
    if name is None:
        return ""
    return str(name).strip().lower().replace(" ", "").replace("_", "")


def _to_float(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s_raw = str(value).strip()
    s = s_raw.replace(" ", "").replace("%", "").replace(",", ".")
    if not s:
        return None
    try:
        v = float(s)
    except ValueError:
        return None
    if "%" in s_raw or abs(v) > 1:
        return v / 100.0
    return v


def _to_percent_points(value: object) -> Optional[float]:
    # Output in percent points: 3.63 means 3.63%
    if value is None:
        return None
    if isinstance(value, (int, float)):
        v = float(value)
    else:
        s_raw = str(value).strip()
        s = s_raw.replace(" ", "").replace("%", "").replace(",", ".")
        if not s:
            return None
        try:
            v = float(s)
        except ValueError:
            return None

    # Normalize multiple possible encodings:
    # 0.0363 -> 3.63 ; 3.63 -> 3.63 ; 363 -> 3.63
    if abs(v) <= 1:
        return v * 100.0
    if abs(v) > 100:
        return v / 100.0
    return v


def _first_word(text: object) -> str:
    s = "" if text is None else str(text).strip()
    if not s:
        return ""
    return s.split()[0].upper()


def _second_word(text: object) -> str:
    s = "" if text is None else str(text).strip()
    parts = s.split()
    if len(parts) < 2:
        return "INCONNU"
    return parts[1].upper()


def _normalize_bank_name(bank: object) -> str:
    b = "" if bank is None else str(bank).strip().upper()
    if b == "SGMB":
        return "SAHAM"
    return b or "INCONNU"


def _safe_sheet_name(name: str) -> str:
    bad = ['\\', '/', '*', '[', ']', ':', '?']
    clean = name
    for ch in bad:
        clean = clean.replace(ch, "_")
    clean = clean.strip() or "SHEET"
    return clean[:31]


def _extract_maturity_label(details: object) -> str:
    s = "" if details is None else str(details).strip().lower()
    if not s:
        return "inconnue"
    # Handle common abbreviations found in source text:
    # sem, s, sem., jr, jrs, j
    s = re.sub(r"\bsem\.\b", "sem", s)
    matches = re.findall(
        r"(\d+)\s*(semaines?|semaine|sem|s|mois|ans?|an|jours?|jrs?|jr|j)\b",
        s,
    )
    if not matches:
        return "inconnue"
    num, unit = matches[-1]
    unit = unit.lower()
    if unit in {"s", "sem"} or unit.startswith("semaine"):
        unit_norm = "semaine" if num == "1" else "semaines"
    elif unit in {"j", "jr", "jrs"} or unit.startswith("jour"):
        unit_norm = "jour" if num == "1" else "jours"
    elif unit.startswith("mois"):
        unit_norm = "mois"
    else:
        unit_norm = "an" if num == "1" else "ans"
    return f"{num} {unit_norm}"


def _maturity_sort_key(label: str) -> tuple[int, int]:
    m = re.match(r"^\s*(\d+)\s*(jour|jours|semaine|semaines|mois|an|ans)\s*$", str(label).lower())
    if not m:
        return (99, 10**9)
    n = int(m.group(1))
    u = m.group(2)
    if u.startswith("jour"):
        return (0, n)
    if u.startswith("semaine"):
        return (1, n)
    if u.startswith("mois"):
        return (2, n)
    return (3, n)


def _find_interest_column(columns: list[str]) -> Optional[str]:
    preferred = ["interestrate", "interest", "intreaste", "interet", "coupon"]
    norm_map = {_norm_col(c): c for c in columns}
    for key in preferred:
        for nc, original in norm_map.items():
            if key in nc:
                return original
    return None


def _find_code_column(columns: list[str]) -> Optional[str]:
    preferred = ["codeisin", "isin", "instrid", "instrumentid", "code"]
    norm_map = {_norm_col(c): c for c in columns}
    for key in preferred:
        for nc, original in norm_map.items():
            if key in nc:
                return original
    return None


@st.cache_resource(show_spinner=False)
def _get_fetcher(cache_dir: str) -> BamCurveFetcher:
    return BamCurveFetcher(cache_dir=cache_dir)


@st.cache_data(ttl=86400, show_spinner=False)
def _get_curve_for_date(curve_date: date, cache_dir: str) -> tuple[list[int], list[float]] | None:
    fetcher = _get_fetcher(cache_dir)
    try:
        return fetcher.get_curve(curve_date)
    except Exception:
        return None


def _compute_taux_bdt(
    df: pd.DataFrame,
    cache_dir: str,
    progress,
    max_workers: int = 8,
) -> tuple[pd.Series, list[date], list[date]]:
    issue_dates = df["ISSUEDT_DT"].dt.date
    unique_dates = sorted({d for d in issue_dates if pd.notna(d)})

    curves: dict[date, tuple[list[int], list[float]] | None] = {}
    total_dates = max(len(unique_dates), 1)

    def _fetch_one(d: date) -> tuple[date, tuple[list[int], list[float]] | None]:
        return d, _get_curve_for_date(d, cache_dir)

    workers = max(1, min(int(max_workers), 12))
    with ThreadPoolExecutor(max_workers=workers) as ex:
        for i, (d, curve) in enumerate(ex.map(_fetch_one, unique_dates), start=1):
            curves[d] = curve
            progress.progress(min(i / total_dates, 1.0), text=f"Chargement des courbes BAM: {i}/{total_dates}")

    ok_dates = [d for d, c in curves.items() if c]
    ko_dates = [d for d, c in curves.items() if not c]

    out = pd.Series(index=df.index, dtype=float)
    total_rows = max(len(df), 1)
    for i, idx in enumerate(df.index, start=1):
        issue_ts = df.at[idx, "ISSUEDT_DT"]
        maturity_ts = df.at[idx, "MATURITYDT_L_DT"]
        if pd.isna(issue_ts) or pd.isna(maturity_ts):
            out.at[idx] = pd.NA
            continue
        maturity_days = int((maturity_ts - issue_ts).days)
        if maturity_days <= 0:
            out.at[idx] = pd.NA
            continue
        curve = curves.get(issue_ts.date())
        if not curve:
            out.at[idx] = pd.NA
            continue
        mt, tx = curve
        try:
            out.at[idx] = float(calcul_taux(maturity_days, mt, tx, issue_ts.date()))
        except Exception:
            out.at[idx] = pd.NA

        progress.progress(min(i / total_rows, 1.0), text=f"Interpolation Taux BDT: {i}/{total_rows}")
    return out, ok_dates, ko_dates


def _style_sheet(ws) -> None:
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    bold = Font(bold=True)
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_col = ws.max_column
    for c in range(1, max_col + 1):
        cell = ws.cell(1, c)
        cell.fill = yellow
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for r in range(2, ws.max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            cell.fill = gray if c <= 5 else yellow
            cell.border = border

    widths = {"A": 18, "B": 38, "C": 18, "D": 18, "E": 18, "F": 12, "G": 12, "H": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _append_spread_summary(ws, df_sheet: pd.DataFrame, data_rows: int) -> None:
    blue = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    gray = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    bold = Font(bold=True)
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    start = data_rows + 3
    headers = ["MATURITE", "MOYENNE SPREAD", "SPREAD MAX", "SPREAD MIN"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(start, i, h)
        c.fill = blue
        c.font = bold
        c.alignment = Alignment(horizontal="center")
        c.border = border

    buckets: dict[str, list[float]] = {}
    tmp = df_sheet.copy()
    tmp["MAT_LABEL"] = tmp["DETAILS DU TITRE"].map(_extract_maturity_label)
    tmp["SPREAD_NUM"] = pd.to_numeric(tmp["Spread"], errors="coerce")
    tmp = tmp.dropna(subset=["SPREAD_NUM"])
    tmp = tmp[tmp["SPREAD_NUM"] >= 0]
    for _, row in tmp.iterrows():
        label = str(row["MAT_LABEL"]).strip()
        if not label or label == "inconnue":
            continue
        buckets.setdefault(label, []).append(float(row["SPREAD_NUM"]))

    rr = start + 1
    for label in sorted(buckets, key=_maturity_sort_key):
        vals = buckets[label]
        avg = sum(vals) / len(vals)
        mx = max(vals)
        mn = min(vals)
        row_vals = [
            label,
            f"{round(avg):.0f} bps",
            f"{round(mx):.0f} bps",
            f"{round(mn):.0f} bps",
        ]
        for i, v in enumerate(row_vals, start=1):
            c = ws.cell(rr, i, v)
            c.fill = gray
            c.border = border
            c.alignment = Alignment(horizontal="center")
            if i == 1:
                c.font = bold
        rr += 1


def _apply_number_formats(ws, header_idx: dict[str, int], data_rows: int) -> None:
    for r in range(2, data_rows + 2):
        if "DATE D'EMISSION" in header_idx:
            ws.cell(r, header_idx["DATE D'EMISSION"]).number_format = "dd/mm/yyyy"
        if "DATE D'ECHEANCE" in header_idx:
            ws.cell(r, header_idx["DATE D'ECHEANCE"]).number_format = "dd/mm/yyyy"
        if "Maturite residuelle" in header_idx:
            ws.cell(r, header_idx["Maturite residuelle"]).number_format = "0.00"
        if "TAUX BDT" in header_idx:
            ws.cell(r, header_idx["TAUX BDT"]).number_format = "0.00%"
        if "Spread" in header_idx:
            ws.cell(r, header_idx["Spread"]).number_format = "0"
        if "TAUX D'INTERET" in header_idx:
            ws.cell(r, header_idx["TAUX D'INTERET"]).number_format = "0.00"


def _make_output_workbook(df_final: pd.DataFrame, instrument_choice: str) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    df_export = df_final.where(pd.notna(df_final), None)
    visible_cols = [
        "CODE",
        "DETAILS DU TITRE",
        "DATE D'EMISSION",
        "DATE D'ECHEANCE",
        "Maturite residuelle",
        "TAUX BDT",
        "Spread",
        "TAUX D'INTERET",
    ]

    # Feuille globale
    ws_all = wb.create_sheet("TOUT")
    ws_all.append(visible_cols)
    for row in df_export[visible_cols].itertuples(index=False, name=None):
        ws_all.append(list(row))
    _style_sheet(ws_all)
    header_idx_all = {str(ws_all.cell(1, c).value): c for c in range(1, ws_all.max_column + 1)}
    _apply_number_formats(ws_all, header_idx_all, len(df_export))
    _append_spread_summary(ws_all, df_export, len(df_export))

    # Feuilles par banque (2eme mot ENGLONGNAME)
    for bank, g in df_export.groupby("BANQUE", sort=True):
        prefix = instrument_choice
        sheet_name = _safe_sheet_name(f"{prefix}_{bank}")
        ws = wb.create_sheet(sheet_name)
        ws.append(visible_cols)
        for row in g[visible_cols].itertuples(index=False, name=None):
            ws.append(list(row))
        _style_sheet(ws)
        header_idx = {str(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
        _apply_number_formats(ws, header_idx, len(g))
        _append_spread_summary(ws, g, len(g))

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _prepare_filtered_data(
    df: pd.DataFrame,
    instrument_choice: str,
    year_start: int,
    year_end: int,
    maturity_min_years: float,
    maturity_max_years: float,
) -> pd.DataFrame:
    dff = df.copy()
    dff["ISSUEDT_DT"] = pd.to_datetime(dff["ISSUEDT"], errors="coerce")
    dff["MATURITYDT_L_DT"] = pd.to_datetime(dff["MATURITYDT_L"], errors="coerce")
    dff["INSTRUMENT_ENGLONGNAME"] = dff["ENGLONGNAME"].map(_first_word)
    dff["BANQUE"] = dff["ENGLONGNAME"].map(_second_word).map(_normalize_bank_name)
    dff["MATURITE"] = (dff["MATURITYDT_L_DT"] - dff["ISSUEDT_DT"]).dt.days
    dff["MATURITE_ANNEE"] = dff["MATURITE"] / 365.0

    if instrument_choice != "TOUT":
        dff = dff[dff["INSTRUMENT_ENGLONGNAME"] == instrument_choice].copy()

    dff = dff[dff["ISSUEDT_DT"].dt.year.between(year_start, year_end, inclusive="both")].copy()
    dff = dff[dff["MATURITE_ANNEE"].between(maturity_min_years, maturity_max_years, inclusive="both")].copy()

    dff = dff.sort_values(["ISSUEDT_DT", "ENGLONGNAME"]).reset_index(drop=True)
    return dff


def main() -> None:
    st.set_page_config(page_title="Suivi OPCVM - ABB", layout="wide", initial_sidebar_state="expanded")
    _inject_theme_css()
    logo_path = _resolve_logo_path()

    if "auth_ok" not in st.session_state:
        st.session_state.auth_ok = False

    if not st.session_state.auth_ok:
        _render_login(logo_path)
        return

    with st.sidebar:
        if logo_path is not None:
            st.image(str(logo_path), width=45)
        st.markdown("## Plateforme de Spread")
        st.markdown("---")
        st.markdown("### Navigation")
        if st.button("Se deconnecter", use_container_width=True):
            st.session_state.auth_ok = False
            st.rerun()
        st.markdown("### Parametres")
        sheet_name = st.text_input("Nom de feuille source", value="TOUT")
        instrument_choice = st.selectbox("Instrument (1er mot ENGLONGNAME)", INSTRUMENTS, index=0)
        year_start = st.number_input("Annee debut ISSUEDT", min_value=1990, max_value=2100, value=2023, step=1)
        year_end = st.number_input("Annee fin ISSUEDT", min_value=1990, max_value=2100, value=2026, step=1)
        maturity_min_years = st.number_input("Maturite min (annees)", min_value=0.0, max_value=100.0, value=0.0, step=0.25)
        maturity_max_years = st.number_input("Maturite max (annees)", min_value=0.0, max_value=100.0, value=10.0, step=0.25)
        max_workers = st.slider("Workers BAM", min_value=2, max_value=12, value=8, step=1)
        cache_dir = st.text_input("Dossier cache BAM", value="cache_bam_curves")

    _render_top_banner(logo_path)
    st.markdown("## Analyse du segment")
    st.markdown(
        "<div class='abb-card'>Chargement du fichier, filtrage CD/BSF, interpolation BAM, export Excel colore.</div>",
        unsafe_allow_html=True,
    )

    if int(year_start) > int(year_end):
        st.error("Annee debut doit etre <= annee fin.")
        return
    if float(maturity_min_years) > float(maturity_max_years):
        st.error("Maturite min doit etre <= maturite max.")
        return

    uploaded = st.file_uploader("Fichier Excel input (.xlsx)", type=["xlsx"])
    if uploaded is None:
        return

    input_bytes = uploaded.getvalue()
    input_name = uploaded.name

    try:
        xls = pd.ExcelFile(BytesIO(input_bytes), engine="openpyxl")
        selected_sheet = sheet_name.strip()
        if selected_sheet not in xls.sheet_names:
            selected_sheet = xls.sheet_names[0]
            st.warning(f"Feuille '{sheet_name}' introuvable. Utilisation de '{selected_sheet}'.")
        df = pd.read_excel(xls, sheet_name=selected_sheet, engine="openpyxl")
        missing = sorted(REQUIRED_COLUMNS - set(df.columns))
        if missing:
            raise ValueError(f"Colonnes manquantes: {missing}")
    except Exception as exc:
        st.error(f"Lecture impossible: {exc}")
        return

    guessed_interest = _find_interest_column(df.columns.tolist())
    interest_options = ["(aucune)"] + df.columns.tolist()
    default_idx = 0 if guessed_interest is None else interest_options.index(guessed_interest)
    chosen_interest = st.selectbox("Colonne Interest", options=interest_options, index=default_idx)
    interest_col = None if chosen_interest == "(aucune)" else chosen_interest

    if st.button("Lancer pricing", type="primary"):
        dff = _prepare_filtered_data(
            df=df,
            instrument_choice=instrument_choice,
            year_start=int(year_start),
            year_end=int(year_end),
            maturity_min_years=float(maturity_min_years),
            maturity_max_years=float(maturity_max_years),
        )
        if dff.empty:
            st.warning("Aucune ligne apres filtres.")
            return

        progress = st.progress(0.0, text="Preparation...")
        taux_bdt, ok_dates, ko_dates = _compute_taux_bdt(
            dff,
            cache_dir=cache_dir,
            progress=progress,
            max_workers=int(max_workers),
        )
        dff["Taux BDT"] = taux_bdt

        if interest_col is not None and interest_col in dff.columns:
            dff["INTERESTRATE_PCT"] = dff[interest_col].map(_to_percent_points)
            # Regle demandee:
            # 3.63 - 3.29 = 0.34 ; 0.01 = 1 bps => 0.34 = 34 bps
            dff["Spread"] = (dff["INTERESTRATE_PCT"] - (dff["Taux BDT"] * 100.0)) * 100.0
            before_spread = len(dff)
            no_spread = int(dff["Spread"].isna().sum())
            lt_10 = int((dff["Spread"] < 10).sum())
            gt_75 = int((dff["Spread"] > 75).sum())
            # Keep only spreads in [10, 75] bps
            dff = dff[dff["Spread"].notna() & (dff["Spread"] >= 10) & (dff["Spread"] <= 75)].copy()
            removed_spread = before_spread - len(dff)
        else:
            dff["Spread"] = pd.NA
            dff["INTERESTRATE_PCT"] = pd.NA
            no_spread = len(dff)
            lt_10 = 0
            gt_75 = 0
            removed_spread = 0

        code_col = _find_code_column(dff.columns.tolist())
        if code_col is None:
            dff["CODE"] = pd.NA
            code_col = "CODE"

        out_df = pd.DataFrame(
            {
                "CODE": dff[code_col],
                "DETAILS DU TITRE": dff["ENGLONGNAME"],
                "DATE D'EMISSION": dff["ISSUEDT_DT"],
                "DATE D'ECHEANCE": dff["MATURITYDT_L_DT"],
                "Maturite residuelle": dff["MATURITE_ANNEE"],
                "TAUX BDT": dff["Taux BDT"],
                "Spread": dff["Spread"],
                "TAUX D'INTERET": dff["INTERESTRATE_PCT"],
                "BANQUE": dff["BANQUE"],
            }
        )

        # Onglets banques utilises sur la colonne BANQUE, mais on retire BANQUE des colonnes visibles exportees.
        export_cols = [
            "CODE",
            "DETAILS DU TITRE",
            "DATE D'EMISSION",
            "DATE D'ECHEANCE",
            "Maturite residuelle",
            "TAUX BDT",
            "Spread",
            "TAUX D'INTERET",
            "BANQUE",
        ]
        out_df = out_df[export_cols].copy()

        try:
            output_bytes = _make_output_workbook(out_df, instrument_choice=instrument_choice)
        except Exception as exc:
            st.error(f"Erreur generation fichier: {exc}")
            return

        filled = int(pd.to_numeric(out_df["TAUX BDT"], errors="coerce").notna().sum())
        st.success(
            f"Termine. Lignes: {len(out_df)} | Taux BDT remplis: {filled} | Onglets banques: {out_df['BANQUE'].nunique()}"
        )
        st.info(f"Dates BAM OK: {len(ok_dates)} | Dates BAM manquantes: {len(ko_dates)}")
        st.write(
            {
                "lignes_initiales_apres_filtres": before_spread if interest_col is not None else len(out_df),
                "lignes_supprimees_spread_total": int(removed_spread),
                "lignes_spread_absent": int(no_spread),
                "lignes_spread_lt_10bps": int(lt_10),
                "lignes_spread_gt_75bps": int(gt_75),
            }
        )
        st.dataframe(out_df.drop(columns=["BANQUE"]).head(30), use_container_width=True)

        out_name = f"{input_name.rsplit('.', 1)[0]}_organise_pricing.xlsx"
        st.download_button(
            label="Telecharger le fichier organise",
            data=output_bytes,
            file_name=out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


if __name__ == "__main__":
    main()
